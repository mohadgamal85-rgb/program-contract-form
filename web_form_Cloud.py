import streamlit as st
from openpyxl import Workbook, load_workbook
from io import BytesIO
import os

st.set_page_config(page_title="Program & Contract Data Entry", page_icon="📋", layout="centered")
st.title("📋 Program & Contract Data Entry Form")

st.info("On the cloud: files don’t persist. Upload an existing Excel (optional), add rows, then **Download** the updated file.")

# --- Load or create workbook in memory ---
uploaded = st.file_uploader("Upload existing Excel (optional):", type=["xlsx"])
if "wb_bytes" not in st.session_state:
    if uploaded:
        # read uploaded workbook into memory
        wb = load_workbook(uploaded)
    else:
        # create a fresh workbook with header
        wb = Workbook()
        sheet = wb.active
        sheet.title = "dataIn"
        sheet.append([
            "Program Name", "Program Code", "Program Budget Value",
            "First Contract Name", "First Contractor Name",
            "Contract Start Date", "Contract Finish Date", "Contract Value"
        ])
    # store workbook bytes in session
    bio = BytesIO()
    wb.save(bio)
    st.session_state.wb_bytes = bio.getvalue()
    st.session_state.filename = "MainData.xlsx"

# Helper to get workbook from session
def get_wb():
    return load_workbook(BytesIO(st.session_state.wb_bytes))

# --- Form fields ---
with st.form("entry_form"):
    col1, col2 = st.columns(2)
    program_name = col1.text_input("Program Name")
    program_code = col2.text_input("Program Code")
    program_budget = col1.number_input("Program Budget Value", min_value=0.0, format="%.2f")
    first_contract_name = col2.text_input("First Contract Name")
    first_contractor_name = col1.text_input("First Contractor Name")
    contract_start_date = col2.date_input("Contract Start Date")
    contract_finish_date = col1.date_input("Contract Finish Date")
    contract_value = col2.number_input("Contract Value", min_value=0.0, format="%.2f")

    submitted = st.form_submit_button("➕ Add Row")
    if submitted:
        # load wb from session, append row, save back
        wb = get_wb()
        sheet = wb["dataIn"] if "dataIn" in wb.sheetnames else wb.active
        sheet.append([
            program_name,
            program_code,
            program_budget,
            first_contract_name,
            first_contractor_name,
            str(contract_start_date),
            str(contract_finish_date),
            contract_value
        ])
        bio = BytesIO()
        wb.save(bio)
        st.session_state.wb_bytes = bio.getvalue()
        st.success("Row added to workbook in memory ✅")

# --- Download latest workbook ---
st.download_button(
    "⬇️ Download Excel (MainData.xlsx)",
    data=st.session_state.wb_bytes,
    file_name=st.session_state.get("filename", "MainData.xlsx"),
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# --- Optional: preview last few rows ---
with st.expander("Preview last 10 rows"):
    wb = get_wb()
    sh = wb["dataIn"]
    data = [[cell.value for cell in row] for row in sh.iter_rows(values_only=False)]
    # simple print preview
    st.write(f"Total rows (including header): {len(data)}")
    st.dataframe(data[-10:] if len(data) > 10 else data)
